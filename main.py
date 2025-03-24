import os
import xml.etree.ElementTree as ET
import openpyxl
import tkinter as tk
from tkinter import filedialog, messagebox
import subprocess

# Mapeamento dos campos do XML do CTe
CAMPOS_CTE = {
    "UF ORIGEM": ".//ide/UFIni",
    "UF DESTINO": ".//ide/UFFim",
    "CNPJ ORIGEM": ".//emit/CNPJ",
    "NUMERO CTE": ".//ide/nCT",
    "DATA DE EMISSÃO DO CTE": ".//ide/dhEmi",
    "PEDIDO DE COMPRA/OC": ".//infCTeNorm/infDoc/infNFe/chave",
    "NUMERO DA NOTA FISCAL (NF)": ".//infCTeNorm/infDoc/infNFe/chave",
    "VALOR MERCADORIA": ".//vPrest/vTPrest",
    "PESO CÁLCULO (KG)": ".//infCTeNorm/infCarga/vCarga",
    "PEDÁGIO": ".//vPrest/Comp[xNome='PEDAGIO']/vComp",
    "FORNECEDOR": ".//emit/xNome",
    "MODAL": ".//ide/modal",
    "VEÍCULO": ".//infModal/rodo/RNTRC",
    "IMPOSTO (%)": ".//imp/ICMS/ICMS00/pICMS",
    "VALOR CTE": ".//vPrest/vRec"
}

def extrair_dados_xml(xml_path):
    """Lê o XML e extrai os dados"""
    tree = ET.parse(xml_path)
    root = tree.getroot()

    dados = {}
    for campo, xpath in CAMPOS_CTE.items():
        elemento = root.find(xpath)
        dados[campo] = elemento.text if elemento is not None else ""

    return dados

def preencher_planilha(dados, excel_path):
    """Insere os dados extraídos na aba 'CONTROLE' sem sobrescrever registros antigos"""
    if not os.path.exists(excel_path):
        messagebox.showerror("Erro", "A planilha selecionada não foi encontrada!")
        return

    # Abrir a planilha
    wb = openpyxl.load_workbook(excel_path)

    if "CONTROLE" not in wb.sheetnames:
        messagebox.showerror("Erro", "A aba 'CONTROLE' não foi encontrada!")
        return

    ws = wb["CONTROLE"]

    # Identificar as colunas existentes
    colunas_existentes = [cell.value for cell in ws[1]]

    # Criar a lista de valores organizados para inserção correta
    valores_ordenados = [dados.get(coluna, "N/A") for coluna in colunas_existentes]

    # Inserir os dados na primeira linha vazia
    linha_vazia = ws.max_row + 1
    for col_idx, valor in enumerate(valores_ordenados, start=1):
        ws.cell(row=linha_vazia, column=col_idx, value=valor)

    # Salvar e fechar a planilha corretamente
    wb.save(excel_path)
    wb.close()

    # Abrir a planilha automaticamente
    subprocess.Popen(["start", excel_path], shell=True)

    messagebox.showinfo("Sucesso", "Os dados foram extraídos e adicionados à planilha!")

def selecionar_xml():
    """Abre o explorador de arquivos para selecionar um XML"""
    caminho_xml = filedialog.askopenfilename(filetypes=[("Arquivos XML", "*.xml")])
    if caminho_xml:
        entry_xml.delete(0, tk.END)
        entry_xml.insert(0, caminho_xml)

def selecionar_planilha():
    """Abre o explorador de arquivos para selecionar uma planilha"""
    caminho_excel = filedialog.askopenfilename(filetypes=[("Planilhas Excel", "*.xlsx")])
    if caminho_excel:
        entry_excel.delete(0, tk.END)
        entry_excel.insert(0, caminho_excel)

def processar():
    """Executa a extração e preenchimento"""
    xml_path = entry_xml.get()
    excel_path = entry_excel.get()

    if xml_path and excel_path:
        try:
            dados = extrair_dados_xml(xml_path)
            if dados:
                preencher_planilha(dados, excel_path)
            else:
                messagebox.showerror("Erro", "Não foi possível extrair os dados do XML.")
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro: {e}")
    else:
        messagebox.showerror("Erro", "Selecione um arquivo XML e uma planilha Excel.")

# Criar a interface gráfica com Tkinter
root = tk.Tk()
root.title("Automação CTe XML para Excel")
root.geometry("400x200")

tk.Label(root, text="Arquivo XML:").pack()
entry_xml = tk.Entry(root, width=50)
entry_xml.pack()
tk.Button(root, text="Selecionar XML", command=selecionar_xml).pack()

tk.Label(root, text="Planilha Excel:").pack()
entry_excel = tk.Entry(root, width=50)
entry_excel.pack()
tk.Button(root, text="Selecionar Planilha", command=selecionar_planilha).pack()

tk.Button(root, text="Processar", command=processar).pack()
tk.Button(root, text="Sair", command=root.quit).pack()

root.mainloop()
